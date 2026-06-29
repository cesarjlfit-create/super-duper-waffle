import json
import io
import traceback
import logging
from contextlib import asynccontextmanager
from typing import Optional

import qrcode
import qrcode.image.pil
from reportlab.lib.pagesizes import A6
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import config
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from supabase import create_client, Client

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

supabase: Optional[Client] = None
gemini_client: Optional[genai.Client] = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global supabase
    try:
        supabase = create_client(config.SUPABASE_URL, config.SUPABASE_KEY)
        logger.info("Serviços iniciados com sucesso.")
    except Exception as e:
        logger.error(f"Erro ao iniciar clientes: {e}")
    yield


app = FastAPI(title="WMS Pulmão API", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_supabase() -> Client:
    if supabase is None:
        raise HTTPException(status_code=503, detail="Banco de dados indisponível.")
    return supabase


def get_gemini() -> genai.Client:
    if gemini_client is None:
        raise HTTPException(status_code=503, detail="Serviço de IA indisponível.")
    return gemini_client


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class BaixaItem(BaseModel):
    id_item: int = Field(..., gt=0)
    quantidade_retirada: int = Field(..., gt=0)


class ItemPalete(BaseModel):
    descricao: str
    codigo_pa: str
    lote: str
    data_validade: str
    quantidade: int = Field(..., gt=0)


class AdicionarPaleteBody(BaseModel):
    id_palete: int = Field(..., gt=0)
    coluna: str
    nivel: int = Field(..., gt=0)
    lado: str
    itens: list[ItemPalete] = Field(..., min_length=1)


class AtualizarQuantidade(BaseModel):
    id_item: int = Field(..., gt=0)
    quantidade: int = Field(..., ge=0)


# ===========================================================================
# SAÚDE
# ===========================================================================
@app.get("/api/healthz", tags=["Saúde"])
async def health_check(db: Client = Depends(get_supabase)):
    try:
        db.table("posicoes_estoque").select("id").limit(1).execute()
        db_status = "Conectado"
    except Exception:
        db_status = "Erro de Conexão"
    return {"status": "ok", "db": db_status}


# ===========================================================================
# PICKING
# ===========================================================================
@app.get("/api/palete/consulta/{id_palete}", tags=["Picking"])
def consultar_palete(id_palete: int, db: Client = Depends(get_supabase)):
    resp_palete = (
        db.table("paletes").select("id_palete, posicao_id").eq("id_palete", id_palete).execute()
    )
    if not resp_palete.data:
        raise HTTPException(status_code=404, detail=f"Palete {id_palete} não encontrado.")

    palete = resp_palete.data[0]
    resp_pos = (
        db.table("posicoes_estoque").select("coluna, nivel, lado")
        .eq("id", palete["posicao_id"]).execute()
    )
    posicao = resp_pos.data[0] if resp_pos.data else {}
    resp_itens = db.table("itens_palete").select("*").eq("palete_id", id_palete).execute()

    return {
        "palete": {
            "id": palete["id_palete"],
            "posicao": (
                f"{posicao.get('coluna','')}{posicao.get('nivel','')} ({posicao.get('lado','')})"
                if posicao else "N/A"
            ),
        },
        "itens": resp_itens.data or [],
    }


@app.post("/api/palete/baixar", tags=["Picking"])
def dar_baixa(body: BaixaItem, db: Client = Depends(get_supabase)):
    resp = db.table("itens_palete").select("*").eq("id", body.id_item).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail="Item não encontrado.")

    item = resp.data[0]
    nova_qtd = item["quantidade"] - body.quantidade_retirada

    if nova_qtd < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Quantidade insuficiente. Disponível: {item['quantidade']}",
        )

    if nova_qtd == 0:
        db.table("itens_palete").delete().eq("id", body.id_item).execute()
        restantes = (
            db.table("itens_palete").select("id").eq("palete_id", item["palete_id"]).execute()
        )
        if not restantes.data:
            db.table("paletes").delete().eq("id_palete", item["palete_id"]).execute()
            return {"mensagem": "Item removido e palete vazio excluído."}
        return {"mensagem": "Item removido (estoque zerado)."}

    db.table("itens_palete").update({"quantidade": nova_qtd}).eq("id", body.id_item).execute()
    return {"mensagem": f"Baixa realizada. Novo saldo: {nova_qtd}"}


@app.patch("/api/palete/item/atualizar-quantidade", tags=["Picking"])
def atualizar_quantidade_item(body: AtualizarQuantidade, db: Client = Depends(get_supabase)):
    resp = db.table("itens_palete").select("id").eq("id", body.id_item).execute()
    if not resp.data:
        raise HTTPException(status_code=404, detail="Item não encontrado.")
    db.table("itens_palete").update({"quantidade": body.quantidade}).eq("id", body.id_item).execute()
    return {"mensagem": f"Quantidade atualizada para {body.quantidade}."}


# ===========================================================================
# MAPA
# ===========================================================================
@app.get("/api/mapa", tags=["Mapa"])
def mapa_armazem(db: Client = Depends(get_supabase)):
    """
    Retorna posições com:
    - num_paletes: contagem de paletes na posição
    - id_palete: ID do primeiro palete (para exibir na célula)
    - disponivel: True se a posição não tiver nenhum palete
    """
    resp_pos = db.table("posicoes_estoque").select("id, coluna, nivel, lado").execute()
    resp_paletes = db.table("paletes").select("posicao_id, id_palete").execute()

    # Mapeia posicao_id → lista de paletes
    mapa: dict[int, list[int]] = {}
    for p in resp_paletes.data or []:
        pos_id = p["posicao_id"]
        mapa.setdefault(pos_id, []).append(p["id_palete"])

    posicoes = []
    for pos in resp_pos.data or []:
        paletes_aqui = mapa.get(pos["id"], [])
        posicoes.append({
            **pos,
            "num_paletes": len(paletes_aqui),
            "id_palete": paletes_aqui[0] if paletes_aqui else None,
            "disponivel": len(paletes_aqui) == 0,
        })

    total = len(posicoes)
    disponiveis = sum(1 for p in posicoes if p["disponivel"])

    return {"posicoes": posicoes, "total": total, "disponiveis": disponiveis}


@app.get("/api/posicao/{posicao_id}/itens", tags=["Mapa"])
def itens_por_posicao(posicao_id: int, db: Client = Depends(get_supabase)):
    resp_paletes = (
        db.table("paletes").select("id_palete").eq("posicao_id", posicao_id).execute()
    )
    resultado = []
    for palete in resp_paletes.data or []:
        id_p = palete["id_palete"]
        resp_itens = db.table("itens_palete").select("*").eq("palete_id", id_p).execute()
        resultado.append({"id_palete": id_p, "itens": resp_itens.data or []})

    return {"paletes": resultado}


# ===========================================================================
# RECEBIMENTO
# ===========================================================================
@app.post("/api/palete/escanear", tags=["Recebimento"])
async def escanear_etiqueta(
    file: UploadFile = File(...),
    ai: genai.Client = Depends(get_gemini),
):
    """Lê a etiqueta com IA e devolve os dados extraídos SEM salvar no banco."""
    try:
        file_bytes = await file.read()
        if not file_bytes:
            raise HTTPException(status_code=400, detail="Arquivo de imagem vazio.")

        prompt = (
            "Analise esta etiqueta logística e extraia as informações em JSON com as chaves: "
            "descricao, codigo_pa, lote, data_validade (formato YYYY-MM-DD), quantidade (inteiro, 0 se não encontrar). "
            "Responda APENAS com o JSON."
        )
        response = ai.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Part.from_bytes(data=file_bytes, mime_type=file.content_type or "image/jpeg"),
                prompt,
            ],
            config=types.GenerateContentConfig(response_mime_type="application/json"),
        )
        raw_text = response.text.strip().replace("```json", "").replace("```", "").strip()
        try:
            dados_ia = json.loads(raw_text)
        except json.JSONDecodeError as e:
            logger.error(f"Resposta inválida da IA: {raw_text}")
            raise HTTPException(status_code=502, detail=f"IA retornou resposta inválida: {e}")

        return dados_ia

    except HTTPException:
        raise
    except Exception:
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Erro ao escanear etiqueta.")


@app.post("/api/palete/adicionar", tags=["Recebimento"])
async def adicionar_palete(
    body: AdicionarPaleteBody,
    db: Client = Depends(get_supabase),
):
    """
    Salva o palete e TODOS os seus itens (suporte a múltiplos produtos por palete).
    O frontend já valida lotes duplicados antes de chamar este endpoint.
    """
    try:
        coluna_limpa = body.coluna.strip().upper()
        lado_limpo = body.lado.strip().capitalize()

        resp_pos = (
            db.table("posicoes_estoque").select("id")
            .ilike("coluna", coluna_limpa)
            .eq("nivel", body.nivel)
            .ilike("lado", f"%{lado_limpo}%")
            .execute()
        )
        if not resp_pos.data:
            raise HTTPException(
                status_code=404,
                detail=f"Posição {coluna_limpa}{body.nivel} ({lado_limpo}) não encontrada.",
            )

        posicao_id_real = resp_pos.data[0]["id"]

        duplicado = db.table("paletes").select("id_palete, posicao_id").eq("id_palete", body.id_palete).execute()
        if duplicado.data:
            pos_id = duplicado.data[0]["posicao_id"]
            resp_pos_dup = (
                db.table("posicoes_estoque")
                .select("coluna, nivel, lado")
                .eq("id", pos_id)
                .execute()
            )
            pos_dup = resp_pos_dup.data[0] if resp_pos_dup.data else {}
            posicao_str = f"{pos_dup.get('coluna','?')}{pos_dup.get('nivel','?')}\ ({pos_dup.get('lado','?')})"
            raise HTTPException(
                status_code=409,
                detail=json.dumps({
                    "mensagem": f"Palete {body.id_palete} já existe em {posicao_str}.",
                    "posicao_id": pos_id,
                    "posicao_str": posicao_str,
                    "id_palete": body.id_palete,
                }),
            )

        db.table("paletes").insert(
            {"id_palete": body.id_palete, "posicao_id": posicao_id_real}
        ).execute()

        rows = [
            {
                "palete_id": body.id_palete,
                "descricao": item.descricao or "N/A",
                "codigo_pa": item.codigo_pa or "N/A",
                "lote": item.lote or "N/A",
                "data_validade": item.data_validade or "2099-01-01",
                "quantidade": item.quantidade,
            }
            for item in body.itens
        ]
        db.table("itens_palete").insert(rows).execute()

        total_un = sum(i.quantidade for i in body.itens)
        logger.info(f"Palete {body.id_palete} — {len(body.itens)} item(ns), {total_un} un.")
        return {
            "status": "Sucesso",
            "mensagem": f"Palete {body.id_palete} salvo — {len(body.itens)} produto(s), {total_un} unidades.",
            "id_palete": body.id_palete,
            "total_itens": len(body.itens),
            "total_unidades": total_un,
        }

    except HTTPException:
        raise
    except Exception:
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Erro interno. Verifique os logs.")



# ===========================================================================
# ADICIONAR ITENS A PALETE EXISTENTE
# ===========================================================================
class AdicionarItensBody(BaseModel):
    itens: list[ItemPalete] = Field(..., min_length=1)


@app.post("/api/palete/{id_palete}/itens", tags=["Recebimento"])
def adicionar_itens_palete(
    id_palete: int,
    body: AdicionarItensBody,
    db: Client = Depends(get_supabase),
):
    """
    Adiciona novos itens a um palete já existente no armazém.
    Usado quando o operador quer incluir produtos em um palete já alocado.
    """
    resp_p = db.table("paletes").select("id_palete, posicao_id").eq("id_palete", id_palete).execute()
    if not resp_p.data:
        raise HTTPException(status_code=404, detail=f"Palete {id_palete} não encontrado.")

    rows = [
        {
            "palete_id": id_palete,
            "descricao": item.descricao or "N/A",
            "codigo_pa": item.codigo_pa or "N/A",
            "lote": item.lote or "N/A",
            "data_validade": item.data_validade or "2099-01-01",
            "quantidade": item.quantidade,
        }
        for item in body.itens
    ]
    db.table("itens_palete").insert(rows).execute()

    total_un = sum(i.quantidade for i in body.itens)
    logger.info(f"Palete {id_palete}: {len(body.itens)} item(ns) adicionado(s), {total_un} un.")
    return {
        "status": "Sucesso",
        "mensagem": f"{len(body.itens)} produto(s) adicionado(s) ao palete {id_palete} ({total_un} unidades).",
        "id_palete": id_palete,
        "total_itens": len(body.itens),
        "total_unidades": total_un,
    }


# ===========================================================================
# INVENTÁRIO
# ===========================================================================
def _get_inventario_enriquecido(db: Client) -> list[dict]:
    resp_itens = db.table("itens_palete").select("*").execute()
    resp_paletes = db.table("paletes").select("id_palete, posicao_id").execute()
    resp_pos = db.table("posicoes_estoque").select("id, coluna, nivel, lado").execute()

    palete_para_posicao: dict[int, int] = {
        p["id_palete"]: p["posicao_id"] for p in (resp_paletes.data or [])
    }
    posicao_label: dict[int, str] = {
        p["id"]: f"{p['coluna']}{p['nivel']} ({p['lado']})"
        for p in (resp_pos.data or [])
    }

    itens = []
    for item in resp_itens.data or []:
        pos_id = palete_para_posicao.get(item.get("palete_id"))
        item["posicao"] = posicao_label.get(pos_id, "N/A") if pos_id else "N/A"
        itens.append(item)
    return itens


@app.get("/api/inventario", tags=["Inventário"])
def listar_inventario(db: Client = Depends(get_supabase)):
    itens = _get_inventario_enriquecido(db)
    return {"itens": itens, "total": len(itens)}


@app.get("/api/inventario/exportar", tags=["Inventário"])
def exportar_inventario_excel(db: Client = Depends(get_supabase)):
    itens = _get_inventario_enriquecido(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventário WMS"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["ID", "Palete", "Posição", "Descrição", "Código PA", "Lote", "Validade", "Quantidade"]
    col_widths = [8, 10, 20, 42, 14, 14, 14, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(itens, start=2):
        row_data = [
            item.get("id"), item.get("palete_id"), item.get("posicao", "N/A"),
            item.get("descricao"), item.get("codigo_pa"), item.get("lote"),
            item.get("data_validade"), item.get("quantidade"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventario_wms.xlsx"},
    )


# ===========================================================================
# ETIQUETA PDF — FIX 2: retorna dados do palete para exibir na tela de sucesso
# ===========================================================================
@app.get("/api/palete/{id_palete}/dados", tags=["Recebimento"])
def dados_palete(id_palete: int, db: Client = Depends(get_supabase)):
    """Retorna dados completos do palete recém-criado (usado na tela de sucesso)."""
    resp_p = db.table("paletes").select("id_palete, posicao_id").eq("id_palete", id_palete).execute()
    if not resp_p.data:
        raise HTTPException(status_code=404, detail=f"Palete {id_palete} não encontrado.")

    palete = resp_p.data[0]
    resp_pos = (
        db.table("posicoes_estoque").select("coluna, nivel, lado")
        .eq("id", palete["posicao_id"]).execute()
    )
    pos = resp_pos.data[0] if resp_pos.data else {}
    posicao_str = f"{pos.get('coluna','?')}{pos.get('nivel','?')} ({pos.get('lado','?')})"

    resp_itens = db.table("itens_palete").select("*").eq("palete_id", id_palete).execute()

    return {
        "id_palete": id_palete,
        "posicao": posicao_str,
        "itens": resp_itens.data or [],
    }


@app.get("/api/palete/{id_palete}/etiqueta", tags=["Recebimento"])
def gerar_etiqueta(id_palete: int, db: Client = Depends(get_supabase)):
    """Gera PDF A6 com QR code para impressão."""
    resp_p = db.table("paletes").select("id_palete, posicao_id").eq("id_palete", id_palete).execute()
    if not resp_p.data:
        raise HTTPException(status_code=404, detail=f"Palete {id_palete} não encontrado.")

    palete = resp_p.data[0]
    resp_pos = (
        db.table("posicoes_estoque").select("coluna, nivel, lado")
        .eq("id", palete["posicao_id"]).execute()
    )
    pos = resp_pos.data[0] if resp_pos.data else {}
    posicao_str = f"{pos.get('coluna','?')}{pos.get('nivel','?')} ({pos.get('lado','?')})"

    resp_itens = db.table("itens_palete").select("*").eq("palete_id", id_palete).execute()
    itens = resp_itens.data or []

    # Gera QR code
    qr = qrcode.QRCode(box_size=4, border=2)
    qr.add_data(str(id_palete))
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format="PNG")
    qr_buf.seek(0)

    buf = io.BytesIO()
    W, H = A6

    c = rl_canvas.Canvas(buf, pagesize=A6)
    c.setTitle(f"Etiqueta Palete {id_palete}")

    # Cabeçalho escuro
    c.setFillColor(colors.HexColor("#1e293b"))
    c.rect(0, H - 28*mm, W, 28*mm, fill=True, stroke=False)
    c.setFillColor(colors.HexColor("#facc15"))
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(W / 2, H - 12*mm, "WMS PULMÃO")
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.white)
    c.drawCentredString(W / 2, H - 19*mm, "ETIQUETA DE PALETE")

    # Nº do palete
    c.setFillColor(colors.HexColor("#1e293b"))
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(W / 2, H - 42*mm, f"#{id_palete}")

    # Posição
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(colors.HexColor("#64748b"))
    c.drawCentredString(W / 2, H - 51*mm, "POSIÇÃO")
    c.setFont("Helvetica-Bold", 14)
    c.setFillColor(colors.HexColor("#0f172a"))
    c.drawCentredString(W / 2, H - 60*mm, posicao_str)

    # QR code centralizado
    qr_size = 32*mm
    c.drawImage(qr_buf, (W - qr_size) / 2, H - 98*mm, width=qr_size, height=qr_size)

    # Linha separadora
    c.setStrokeColor(colors.HexColor("#e2e8f0"))
    c.setLineWidth(0.5)
    c.line(8*mm, H - 101*mm, W - 8*mm, H - 101*mm)

    # Lista de itens
    y = H - 108*mm
    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(colors.HexColor("#64748b"))
    c.drawString(8*mm, y, "PRODUTOS NO PALETE:")
    y -= 5*mm

    for item in itens[:6]:
        desc = (item.get("descricao") or "")[:38]
        qtd = item.get("quantidade", 0)
        lote = (item.get("lote") or "")[:10]
        c.setFont("Helvetica-Bold", 7)
        c.setFillColor(colors.HexColor("#0f172a"))
        c.drawString(8*mm, y, f"• {desc}")
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#475569"))
        c.drawString(8*mm, y - 4*mm, f"   Lote: {lote}   Qtd: {qtd} un")
        y -= 10*mm
        if y < 8*mm:
            break

    if len(itens) > 6:
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#94a3b8"))
        c.drawCentredString(W / 2, y, f"... e mais {len(itens) - 6} item(ns)")

    c.save()
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=etiqueta_palete_{id_palete}.pdf"},
    )


# ===========================================================================
# SEED
# ===========================================================================
@app.post("/api/admin/seed-posicoes", tags=["Admin"])
def seed_posicoes(db: Client = Depends(get_supabase)):
    existentes = db.table("posicoes_estoque").select("id").limit(1).execute()
    if existentes.data:
        total = db.table("posicoes_estoque").select("id", count="exact").execute().count
        return {"mensagem": f"Posições já cadastradas. Total: {total}"}

    LADOS = ["Esquerdo", "Direito"]
    COLUNAS = ["A", "B", "C", "D", "E", "F", "G"]
    NIVEIS = [1, 2, 3, 4, 5, 6]

    posicoes = [
        {"coluna": col, "nivel": nv, "lado": lado}
        for lado in LADOS for col in COLUNAS for nv in NIVEIS
    ]
    db.table("posicoes_estoque").insert(posicoes).execute()
    return {"mensagem": f"{len(posicoes)} posições criadas.", "total": len(posicoes)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)
